import io
from openpyxl import load_workbook
from services.openai_extractor import SalesOrderData, SalesInvoiceData, GDNData

def populate_template(
    sales_orders: list[SalesOrderData],
    invoices: list[SalesInvoiceData],
    gdns: list[GDNData],
) -> bytes:
    template_path = "templates/revenue_audit_3tab.xlsx"
    wb = load_workbook(template_path)
    
    def set_cell(sheet, row, col, value):
        if value is not None:
            sheet.cell(row=row, column=col).value = value

    # Sales Order
    so_sheet = wb["Sales Order"]
    if len(sales_orders) > 10:
        print(f"[ExcelExport] Truncating SalesOrders to 10 rows (got {len(sales_orders)})")
    
    for i, so in enumerate(sales_orders[:10], start=3):
        first_item = so.line_items[0] if so.line_items else None
        
        customer = so.customer_name
        if customer is None and first_item:
            customer = first_item.description
            
        qty = so.total_quantity
        if qty is None and first_item:
            qty = first_item.quantity
            
        rate = first_item.rate if first_item else None
        
        set_cell(so_sheet, i, 2, so.so_number)        # B
        set_cell(so_sheet, i, 3, customer)            # C
        set_cell(so_sheet, i, 4, qty)                 # D
        set_cell(so_sheet, i, 5, rate)                # E
        set_cell(so_sheet, i, 7, so.notes)            # G

    # Sales Invoice
    inv_sheet = wb["Sales Invoice"]
    if len(invoices) > 10:
        print(f"[ExcelExport] Truncating SalesInvoices to 10 rows (got {len(invoices)})")
        
    for i, inv in enumerate(invoices[:10], start=3):
        first_item = inv.line_items[0] if inv.line_items else None
        
        qty = inv.total_quantity
        if qty is None and first_item:
            qty = first_item.quantity
            
        rate = first_item.rate if first_item else None
        
        set_cell(inv_sheet, i, 2, inv.customer_name)  # B
        set_cell(inv_sheet, i, 3, inv.invoice_number) # C
        set_cell(inv_sheet, i, 4, inv.invoice_date)   # D
        set_cell(inv_sheet, i, 5, qty)                # E
        set_cell(inv_sheet, i, 6, rate)               # F
        set_cell(inv_sheet, i, 8, inv.total_amount)   # H
        set_cell(inv_sheet, i, 10, inv.notes)         # J

    # GDN
    gdn_sheet = wb["GDN"]
    if len(gdns) > 10:
        print(f"[ExcelExport] Truncating GDNs to 10 rows (got {len(gdns)})")
        
    for i, gdn in enumerate(gdns[:10], start=3):
        first_item = gdn.line_items[0] if gdn.line_items else None
        
        qty = gdn.total_quantity_delivered
        if qty is None and first_item:
            qty = first_item.quantity_delivered
            
        set_cell(gdn_sheet, i, 2, gdn.customer_name)      # B
        set_cell(gdn_sheet, i, 3, gdn.delivered_date)     # C
        set_cell(gdn_sheet, i, 4, qty)                    # D
        set_cell(gdn_sheet, i, 5, gdn.gdn_reference)      # E
        set_cell(gdn_sheet, i, 6, gdn.notes)              # F

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
